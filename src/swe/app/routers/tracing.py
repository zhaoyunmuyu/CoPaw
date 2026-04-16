# -*- coding: utf-8 -*-
# pylint: disable=too-many-branches too-many-statements too-many-locals
"""Tracing API router.

Provides REST API endpoints for tracing analytics.
"""
from datetime import datetime, timedelta
from typing import Optional
import io
import csv
import logging

from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse

from ...tracing import get_trace_manager, has_trace_manager
from ...tracing.models import (
    OverviewStats,
    SkillToolsStats,
    TraceDetail,
    TraceDetailWithTimeline,
    SessionStats,
    UserStats,
)

router = APIRouter(prefix="/tracing", tags=["tracing"])
logger = logging.getLogger(__name__)

_USER_MESSAGE_EXPORT_HEADERS = [
    "trace_id",
    "user_id",
    "session_id",
    "channel",
    "user_message",
    "input_tokens",
    "output_tokens",
    "model_name",
    "start_time",
    "duration_ms",
]


def _parse_date(
    date_str: Optional[str],
    field_name: str,
    add_day: bool = False,
) -> Optional[datetime]:
    """Parse a date string to datetime.

    Args:
        date_str: Date string in YYYY-MM-DD format
        field_name: Field name for error message
        add_day: Whether to add one day to include the end date

    Returns:
        Parsed datetime or None

    Raises:
        HTTPException: If date format is invalid
    """
    if not date_str:
        return None
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        if add_day:
            dt = dt + timedelta(days=1)
        return dt
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid {field_name} format",
        ) from exc


def _check_tracing_available() -> None:
    """Check if tracing is available, raise HTTPException if not."""
    if not has_trace_manager():
        raise HTTPException(
            status_code=503,
            detail="Tracing not available. Enable tracing in configuration.",
        )


def _get_trace_store_or_503():
    """Return the trace store or raise 503 if tracing is unavailable."""
    try:
        manager = get_trace_manager()
        return manager.store
    except RuntimeError as exc:
        raise HTTPException(
            status_code=503,
            detail="Tracing not available",
        ) from exc


def _build_download_headers(filename: str) -> dict[str, str]:
    """Return standard attachment headers for exports."""
    return {"Content-Disposition": f"attachment; filename={filename}"}


def _build_user_message_export_row(message) -> list[object]:
    """Convert a user-message record into an export row."""
    return [
        message.trace_id,
        message.user_id,
        message.session_id,
        message.channel,
        message.user_message or "",
        message.input_tokens,
        message.output_tokens,
        message.model_name or "",
        message.start_time.isoformat() if message.start_time else "",
        message.duration_ms or "",
    ]


def _build_user_messages_json_response(
    messages,
    timestamp: str,
) -> StreamingResponse:
    """Build a JSON export response."""
    import json

    data = [message.model_dump() for message in messages]
    for item in data:
        if item.get("start_time"):
            item["start_time"] = item["start_time"].isoformat()
    content = json.dumps(data, ensure_ascii=False, indent=2)
    return StreamingResponse(
        iter([content]),
        media_type="application/json",
        headers=_build_download_headers(
            f"user_messages_{timestamp}.json",
        ),
    )


def _build_user_messages_xlsx_response(
    messages,
    timestamp: str,
) -> StreamingResponse:
    """Build an XLSX export response."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(
            status_code=400,
            detail="openpyxl not installed. Use csv or json format.",
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "User Messages"

    header_fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid",
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    excel_headers = [
        "Trace ID",
        "User ID",
        "Session ID",
        "Channel",
        "User Message",
        "Input Tokens",
        "Output Tokens",
        "Model Name",
        "Start Time",
        "Duration (ms)",
    ]
    for column, header in enumerate(excel_headers, 1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, message in enumerate(messages, 2):
        for column, value in enumerate(
            _build_user_message_export_row(message),
            1,
        ):
            cell = ws.cell(row=row, column=column, value=value)
            cell.border = thin_border
            if column == 5:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column, width in enumerate(
        [36, 20, 36, 15, 60, 12, 12, 25, 22, 12],
        1,
    ):
        ws.column_dimensions[get_column_letter(column)].width = width

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return StreamingResponse(
        iter([excel_buffer.getvalue()]),
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers=_build_download_headers(
            f"user_messages_{timestamp}.xlsx",
        ),
    )


def _build_user_messages_csv_response(
    messages,
    timestamp: str,
) -> StreamingResponse:
    """Build a CSV export response."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_USER_MESSAGE_EXPORT_HEADERS)
    for message in messages:
        writer.writerow(_build_user_message_export_row(message))
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers=_build_download_headers(
            f"user_messages_{timestamp}.csv",
        ),
    )


@router.get("/overview", response_model=OverviewStats)
async def get_overview(
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> OverviewStats:
    """Get overview statistics for the dashboard.

    Args:
        start_date: Optional start date filter
        end_date: Optional end date filter

    Returns:
        Overview statistics including user counts, token usage,
        model distribution.
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        # Tracing not initialized, return empty stats
        return OverviewStats()

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    return await store.get_overview_stats(start, end)


@router.get("/users", response_model=dict)
async def get_users(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Page size"),
    user_id: Optional[str] = Query(
        None,
        description="Filter by user ID (partial match)",
    ),
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> dict:
    """Get list of users with their statistics.

    Args:
        page: Page number
        page_size: Page size
        user_id: Filter by user ID
        start_date: Start date filter
        end_date: End date filter

    Returns:
        Paginated list of users with stats
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return {"items": [], "total": 0, "page": page, "page_size": page_size}

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    users, total = await store.get_users(
        page,
        page_size,
        user_id,
        start,
        end,
    )
    return {
        "items": [u.model_dump() for u in users],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/users/{user_id}", response_model=UserStats)
async def get_user_stats(
    user_id: str,
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> UserStats:
    """Get statistics for a specific user.

    Args:
        user_id: User identifier
        start_date: Optional start date filter
        end_date: Optional end date filter

    Returns:
        User statistics
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return UserStats(user_id=user_id)

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    return await store.get_user_stats(user_id, start, end)


@router.get("/traces", response_model=dict)
async def get_traces(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Page size"),
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    session_id: Optional[str] = Query(
        None,
        description="Filter by session ID",
    ),
    status: Optional[str] = Query(None, description="Filter by status"),
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> dict:
    """Get list of traces.

    Args:
        page: Page number
        page_size: Page size
        user_id: Filter by user ID
        session_id: Filter by session ID
        status: Filter by status (running, completed, error, cancelled)
        start_date: Start date filter
        end_date: End date filter

    Returns:
        Paginated list of traces
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return {"items": [], "total": 0, "page": page, "page_size": page_size}

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    traces, total = await store.get_traces(
        page=page,
        page_size=page_size,
        user_id=user_id,
        session_id=session_id,
        status=status,
        start_date=start,
        end_date=end,
    )
    return {
        "items": [t.model_dump() for t in traces],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/traces/{trace_id}", response_model=TraceDetail)
async def get_trace_detail(trace_id: str) -> TraceDetail:
    """Get detailed trace with spans.

    Args:
        trace_id: Trace identifier

    Returns:
        Trace detail with all spans

    Raises:
        HTTPException: If trace not found
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError as exc:
        raise HTTPException(
            status_code=503,
            detail="Tracing not available",
        ) from exc

    detail = await store.get_trace_detail(trace_id)
    if detail is None:
        raise HTTPException(status_code=404, detail="Trace not found")

    return detail


@router.get("/models", response_model=dict)
async def get_model_usage(
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> dict:
    """Get model usage statistics.

    Args:
        start_date: Start date filter
        end_date: End date filter

    Returns:
        Model usage statistics
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return {"models": []}

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    stats = await store.get_overview_stats(start, end)
    return {"models": [m.model_dump() for m in stats.model_distribution]}


@router.get("/tools", response_model=dict)
async def get_tool_usage(
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> dict:
    """Get tool usage statistics.

    Args:
        start_date: Start date filter
        end_date: End date filter

    Returns:
        Tool usage statistics
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return {"tools": []}

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    stats = await store.get_overview_stats(start, end)
    return {"tools": [t.model_dump() for t in stats.top_tools]}


@router.get("/skills", response_model=dict)
async def get_skill_usage(
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> dict:
    """Get skill usage statistics.

    Args:
        start_date: Start date filter
        end_date: End date filter

    Returns:
        Skill usage statistics
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return {"skills": []}

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    stats = await store.get_overview_stats(start, end)
    return {"skills": [s.model_dump() for s in stats.top_skills]}


@router.get("/mcp", response_model=dict)
async def get_mcp_usage(
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> dict:
    """Get MCP tool and server usage statistics.

    Args:
        start_date: Start date filter
        end_date: End date filter

    Returns:
        MCP usage statistics
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return {"mcp_tools": [], "mcp_servers": []}

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    stats = await store.get_overview_stats(start, end)
    return {
        "mcp_tools": [t.model_dump() for t in stats.top_mcp_tools],
        "mcp_servers": [s.model_dump() for s in stats.mcp_servers],
    }


@router.get("/sessions", response_model=dict)
async def get_sessions(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Page size"),
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    session_id: Optional[str] = Query(
        None,
        description="Filter by session ID (partial match)",
    ),
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> dict:
    """Get list of sessions with their statistics.

    Args:
        page: Page number
        page_size: Page size
        user_id: Filter by user ID
        session_id: Filter by session ID (partial match)
        start_date: Start date filter
        end_date: End date filter

    Returns:
        Paginated list of sessions with stats
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return {"items": [], "total": 0, "page": page, "page_size": page_size}

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    sessions, total = await store.get_sessions(
        page=page,
        page_size=page_size,
        user_id=user_id,
        session_id=session_id,
        start_date=start,
        end_date=end,
    )
    return {
        "items": [s.model_dump() for s in sessions],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/sessions/{session_id:path}", response_model=SessionStats)
async def get_session_stats(
    session_id: str,
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> SessionStats:
    """Get statistics for a specific session.

    Args:
        session_id: Session identifier
        start_date: Optional start date filter
        end_date: Optional end date filter

    Returns:
        Session statistics
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return SessionStats(session_id=session_id, user_id="", channel="")

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    return await store.get_session_stats(session_id, start, end)


@router.get("/user-messages", response_model=dict)
async def get_user_messages(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Page size"),
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    session_id: Optional[str] = Query(
        None,
        description="Filter by session ID",
    ),
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    query: Optional[str] = Query(
        None,
        description="Search in user message content",
    ),
) -> dict:
    """Get user messages with token info for cost analysis.

    Args:
        page: Page number
        page_size: Page size
        user_id: Filter by user ID
        session_id: Filter by session ID
        start_date: Start date filter
        end_date: End date filter
        query: Search in user message content (partial match)

    Returns:
        Paginated list of user messages with token usage
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return {"items": [], "total": 0, "page": page, "page_size": page_size}

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    messages, total = await store.get_user_messages(
        page=page,
        page_size=page_size,
        user_id=user_id,
        session_id=session_id,
        start_date=start,
        end_date=end,
        query=query,
        export=False,
    )
    return {
        "items": [m.model_dump() for m in messages],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/user-messages/export")
async def export_user_messages(
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    session_id: Optional[str] = Query(
        None,
        description="Filter by session ID",
    ),
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    query: Optional[str] = Query(
        None,
        description="Search in user message content",
    ),
    export_format: str = Query(
        "csv",
        description="Export format: csv, json or xlsx",
        alias="format",
    ),
) -> StreamingResponse:
    """Export user messages with token info for cost analysis.

    Args:
        user_id: Filter by user ID
        session_id: Filter by session ID
        start_date: Start date filter
        end_date: End date filter
        query: Search in user message content (partial match)
        export_format: Export format (csv, json or xlsx)

    Returns:
        StreamingResponse with exported data
    """
    store = _get_trace_store_or_503()

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    messages, _ = await store.get_user_messages(
        page=1,
        page_size=1,  # Ignored in export mode
        user_id=user_id,
        session_id=session_id,
        start_date=start,
        end_date=end,
        query=query,
        export=True,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if export_format == "json":
        return _build_user_messages_json_response(messages, timestamp)
    if export_format == "xlsx":
        return _build_user_messages_xlsx_response(messages, timestamp)
    return _build_user_messages_csv_response(messages, timestamp)


@router.get(
    "/traces/{trace_id}/timeline",
    response_model=TraceDetailWithTimeline,
)
async def get_trace_timeline(trace_id: str) -> TraceDetailWithTimeline:
    """Get trace detail with hierarchical timeline.

    Returns a hierarchical timeline where skill invocations
    are parent nodes containing their tool calls as children.

    Args:
        trace_id: Trace identifier

    Returns:
        Trace detail with hierarchical timeline

    Raises:
        HTTPException: If trace not found
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError as exc:
        raise HTTPException(
            status_code=503,
            detail="Tracing not available",
        ) from exc

    detail = await store.get_trace_detail_with_timeline(trace_id)
    if detail is None:
        raise HTTPException(status_code=404, detail="Trace not found")

    return detail


@router.get("/skills/{skill_name}/tools", response_model=SkillToolsStats)
async def get_skill_tools_stats(
    skill_name: str,
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> SkillToolsStats:
    """Get tools used by a specific skill.

    Returns statistics about which tools are used by the skill,
    including MCP tools and attribution confidence.

    Args:
        skill_name: Skill identifier
        start_date: Start date filter
        end_date: End date filter

    Returns:
        Skill tools statistics
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return SkillToolsStats(skill_name=skill_name)

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    return await store.get_skill_tools_stats(skill_name, start, end)


@router.get("/skills/attribution", response_model=dict)
async def get_skill_attribution(
    tool_name: Optional[str] = Query(
        None,
        description="Filter by tool name",
    ),
    start_date: Optional[str] = Query(
        None,
        description="Start date (YYYY-MM-DD)",
    ),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
) -> dict:
    """Get skill-tool attribution details.

    Returns how tools are attributed to skills, including
    multi-skill attribution weights and confidence levels.

    Args:
        tool_name: Optional filter by tool name
        start_date: Start date filter
        end_date: End date filter

    Returns:
        Tool attribution details
    """
    try:
        manager = get_trace_manager()
        store = manager.store
    except RuntimeError:
        return {"attributions": []}

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date", add_day=True)

    attributions = await store.get_tool_skill_attributions(
        tool_name=tool_name,
        start_date=start,
        end_date=end,
    )
    return {
        "attributions": [a.model_dump() for a in attributions],
    }
