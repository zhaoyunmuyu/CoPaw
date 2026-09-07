# -*- coding: utf-8 -*-
"""Export service for cron data.

Provides methods to export job definitions and execution history to Excel.
"""

import logging
from typing import Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ...models.cron import CronJobModel, ExecutionModel

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting cron data to Excel."""

    def __init__(self) -> None:
        """Initialize export service."""
        pass

    def export_jobs(self, jobs: List[CronJobModel]) -> bytes:
        """Export job definitions to Excel.

        Args:
            jobs: List of job models

        Returns:
            Excel file bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "定时任务定义"

        # Define headers
        headers = [
            ("任务ID", 30),
            ("任务名称", 25),
            ("租户ID", 20),
            ("分行号", 20),
            ("来源标识", 20),
            ("状态", 10),
            ("是否启用", 10),
            ("任务类型", 10),
            ("已执行次数", 12),
            ("Cron表达式", 20),
            ("时区", 12),
            ("分发渠道", 15),
            ("创建者ID", 20),
            ("创建时间", 20),
            ("更新时间", 20),
        ]

        # Set headers with style
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col)].width = width

        # Set data rows
        for row, job in enumerate(jobs, 2):
            ws.cell(row=row, column=1, value=job.id)
            ws.cell(row=row, column=2, value=job.name)
            ws.cell(row=row, column=3, value=job.tenant_id)
            ws.cell(row=row, column=4, value=job.bbk_id)
            ws.cell(row=row, column=5, value=job.source_id)
            ws.cell(row=row, column=6, value=job.status)
            ws.cell(row=row, column=7, value="是" if job.enabled else "否")
            ws.cell(row=row, column=8, value=job.task_type)
            ws.cell(row=row, column=9, value=job.execution_count or 0)
            ws.cell(row=row, column=10, value=job.cron_expr)
            ws.cell(row=row, column=11, value=job.timezone)
            ws.cell(row=row, column=12, value=job.channel)
            ws.cell(row=row, column=13, value=job.creator_user_id)
            ws.cell(
                row=row,
                column=14,
                value=(
                    job.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    if job.created_at
                    else ""
                ),
            )
            ws.cell(
                row=row,
                column=15,
                value=(
                    job.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                    if job.updated_at
                    else ""
                ),
            )

        # Save to bytes
        return self._save_to_bytes(wb)

    def export_executions(self, executions: List[ExecutionModel]) -> bytes:
        """Export execution history to Excel.

        Args:
            executions: List of execution models

        Returns:
            Excel file bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "定时任务执行历史"

        # Define headers
        headers = [
            ("记录ID", 12),
            ("任务ID", 30),
            ("任务名称", 25),
            ("租户ID", 20),
            ("执行状态", 12),
            ("计划时间", 20),
            ("实际开始时间", 20),
            ("结束时间", 20),
            ("耗时(ms)", 12),
            ("错误信息", 40),
            ("是否手动", 10),
            ("Trace ID", 30),
            ("Session ID", 30),
            ("输出预览", 30),
        ]

        # Set headers with style
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col)].width = width

        # Status translation
        status_map = {
            "success": "成功",
            "error": "失败",
            "cancelled": "取消",
            "timeout": "超时",
            "skipped": "跳过",
            "running": "运行中",
        }

        # Set data rows
        for row, exec in enumerate(executions, 2):
            ws.cell(row=row, column=1, value=exec.id)
            ws.cell(row=row, column=2, value=exec.job_id)
            ws.cell(row=row, column=3, value=exec.job_name)
            ws.cell(row=row, column=4, value=exec.tenant_id)
            ws.cell(
                row=row,
                column=5,
                value=status_map.get(exec.status, exec.status),
            )
            ws.cell(
                row=row,
                column=6,
                value=(
                    exec.scheduled_time.strftime("%Y-%m-%d %H:%M:%S")
                    if exec.scheduled_time
                    else ""
                ),
            )
            ws.cell(
                row=row,
                column=7,
                value=(
                    exec.actual_time.strftime("%Y-%m-%d %H:%M:%S")
                    if exec.actual_time
                    else ""
                ),
            )
            ws.cell(
                row=row,
                column=8,
                value=(
                    exec.end_time.strftime("%Y-%m-%d %H:%M:%S")
                    if exec.end_time
                    else ""
                ),
            )
            ws.cell(row=row, column=9, value=exec.duration_ms)
            ws.cell(
                row=row,
                column=10,
                value=exec.error_message[:200] if exec.error_message else "",
            )
            ws.cell(row=row, column=11, value="是" if exec.is_manual else "否")
            ws.cell(row=row, column=12, value=exec.trace_id)
            ws.cell(row=row, column=13, value=exec.session_id)
            ws.cell(row=row, column=14, value=exec.output_preview)

        # Save to bytes
        return self._save_to_bytes(wb)

    @staticmethod
    def _mask_customer_name(name: Any) -> str:
        value = str(name or "")
        if not value or "*" in value or len(value) == 1:
            return value
        if len(value) == 2:
            return f"{value[0]}*"
        return f"{value[0]}{'*' * (len(value) - 2)}{value[-1]}"

    @staticmethod
    def _execution_status(status: Any, async_status: Any) -> str:
        if status == "success" and async_status == "error":
            return "子任务执行失败"
        return {
            "success": "成功",
            "error": "失败",
            "cancelled": "取消",
            "timeout": "超时",
            "skipped": "跳过",
            "running": "运行中",
        }.get(str(status or ""), str(status or "未知"))

    def export_skill_usage_details(self, rows: List[dict[str, Any]]) -> bytes:
        """Export overview execution/customer details to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "客户经理技能明细"
        headers = self._skill_usage_headers()
        self._write_skill_usage_header(ws, headers)
        for row_number, row in enumerate(rows, 2):
            for column, value in enumerate(
                self._skill_usage_row_values(row),
                1,
            ):
                ws.cell(row=row_number, column=column, value=value)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
        )
        return self._save_to_bytes(wb)

    @staticmethod
    def _skill_usage_headers() -> list[tuple[str, int]]:
        return [
            ("定时任务名称", 28),
            ("租户ID", 20),
            ("租户姓名", 18),
            ("分行号", 14),
            ("任务状态", 12),
            ("执行记录ID", 14),
            ("实际执行时间", 20),
            ("执行状态", 16),
            ("是否已读", 10),
            ("定时任务推荐的客户UID", 24),
            ("客户脱敏姓名", 18),
            ("是否查看方案", 14),
            ("是否点击去电访", 14),
            ("是否点击去洞察", 14),
        ]

    @staticmethod
    def _write_skill_usage_header(
        ws: Any,
        headers: list[tuple[str, int]],
    ) -> None:
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        for column, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(column)].width = width

    def _skill_usage_row_values(self, row: dict[str, Any]) -> list[Any]:
        actual_time = row.get("actual_time")
        if actual_time and hasattr(actual_time, "strftime"):
            actual_time = actual_time.strftime("%Y-%m-%d %H:%M:%S")
        return [
            row.get("task_name") or "",
            row.get("tenant_id") or "",
            row.get("tenant_name") or "",
            row.get("bbk_id") or "",
            row.get("task_status") or "",
            row.get("execution_id"),
            actual_time or "",
            self._execution_status(
                row.get("execution_status"),
                row.get("async_status"),
            ),
            "是" if row.get("is_read") else "否",
            row.get("custuid") or "",
            self._mask_customer_name(row.get("cust_nm")),
            "是" if row.get("clicked_plan") else "否",
            "是" if row.get("clicked_phone") else "否",
            "是" if row.get("clicked_insight") else "否",
        ]

    def _save_to_bytes(self, wb: Workbook) -> bytes:
        """Save workbook to bytes.

        Args:
            wb: Workbook object

        Returns:
            Excel file bytes
        """
        from io import BytesIO

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()


# Global export service instance
_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    """Get the export service instance.

    Returns:
        ExportService instance
    """
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
